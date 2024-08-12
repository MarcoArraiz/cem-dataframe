import os
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import tkinter as tk
from tkinter import Listbox, Scrollbar, EXTENDED, Button


def natural_sort_key(s, _nsre=re.compile("([0-9]+)")):
    return [
        int(text) if text.isdigit() else text.lower() for text in re.split(_nsre, s)
    ]


def extract_meters(folder_name):
    parts = folder_name.split(" - ")
    m_start = parts[0].rstrip("m").strip()
    m_end = parts[1].rstrip("m").strip()
    return m_start, m_end


def process_files_in_folder(folder, m_start, m_end):
    minerals_row = {"Hole ID": "", "From": int(float(m_start)), "To": int(float(m_end))}
    elements_row = {"Hole ID": "", "From": int(float(m_start)), "To": int(float(m_end))}
    processed_files = False
    for file in os.listdir(folder):
        if "Summary" in file and file.endswith(".csv"):
            df = pd.read_csv(os.path.join(folder, file))
            if not df.empty:
                minerals_row.update(process_mineral_data(df, m_start, m_end))
                elements_row.update(process_element_data(df))
                processed_files = True

    if not processed_files:
        # Reportar espacio en blanco si no se encuentran archivos
        minerals_row.update(
            {col: None for col in minerals_row if col not in ["Hole ID", "From", "To"]}
        )
        elements_row.update(
            {col: None for col in elements_row if col not in ["Hole ID", "From", "To"]}
        )
    return minerals_row, elements_row


def process_mineral_data(df, m_start, m_end):
    row = {}
    mineral_presence = {}
    total_presence = df["Presence (%)"].sum()
    if total_presence > 0:
        for _, mineral in df.iterrows():
            base_name = re.sub(r"\d+$", "", mineral["Class"]).strip()
            presence = (
                mineral["Presence (%)"] if pd.notna(mineral["Presence (%)"]) else 0
            )
            if presence > 0:
                if base_name not in mineral_presence:
                    mineral_presence[base_name] = 0
                mineral_presence[base_name] += presence

        sorted_minerals = sorted(mineral_presence.keys())
        for mineral in sorted_minerals:
            row[mineral] = mineral_presence[mineral]
    return row


def process_element_data(df):
    row = {}
    elements_to_convert = {"Mo", "Zn", "Ba", "Ag", "As", "Pb"}  # Elementos específicos
    element_data = df[df["Class"] == "TOTAL"]
    if not element_data.empty:
        for col in element_data.columns:
            if col not in ["Class", "Presence (%)", "Density", "Hardness"]:
                value = element_data.iloc[0][col]
                row[col] = value  # Mantener el valor original en la columna principal
                if col in elements_to_convert and value > 0:
                    row[col + " (ppm)"] = int(value * 10000)  # Convertir a PPM y luego a entero
    return row


def format_worksheet(ws):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar formato de número entero a las columnas PPM
            if ws.cell(row=1, column=cell.column).value.endswith("(ppm)"):
                cell.number_format = "0"  # Formato de número entero
                if cell.row > 1:  # Aplicar formato condicional a datos (no a la cabecera)
                    ws.conditional_formatting.add(
                        f"{get_column_letter(cell.column)}2:{get_column_letter(cell.column)}{ws.max_row}",
                        ColorScaleRule(
                            start_type="num",
                            start_value=0,
                            start_color="FFFFFF",
                            end_type="num",
                            end_value=10000,
                            end_color="FF0000",
                        ),
                    )
            elif cell.column >= 4:
                cell.number_format = "0.000"
                if cell.row > 1:  # Aplicar formato condicional para otros elementos
                    ws.conditional_formatting.add(
                        f"{get_column_letter(cell.column)}2:{get_column_letter(cell.column)}{ws.max_row}",
                        ColorScaleRule(
                            start_type="num",
                            start_value=0,
                            start_color="FFFFFF",
                            end_type="num",
                            end_value=100,
                            end_color="00FF00",  # Color verde
                        ),
                    )

    # Set column width for all columns
    for col in range(1, ws.max_column + 1):
        if col < 4:
            ws.column_dimensions[get_column_letter(col)].width = 10
        else:
            ws.column_dimensions[get_column_letter(col)].width = 16


def main():
    root = tk.Tk()
    root.title("Folder Selection")
    current_dir = os.path.abspath(".")
    listbox = Listbox(root, selectmode=EXTENDED, width=100, height=20)
    listbox.grid(row=0, column=0, columnspan=2, padx=10, pady=10)
    scrollbar = Scrollbar(root, orient="vertical", command=listbox.yview)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.grid(row=0, column=2, sticky="ns")

    def update_listbox(directory):
        listbox.delete(0, tk.END)
        items = [
            item
            for item in os.listdir(directory)
            if os.path.isdir(os.path.join(directory, item))
        ]
        items.sort(key=natural_sort_key)
        for item in items:
            listbox.insert(tk.END, item)

    def change_directory():
        selection = listbox.curselection()
        if selection:
            folder = listbox.get(selection[0])
            nonlocal current_dir
            current_dir = os.path.join(current_dir, folder)
            update_listbox(current_dir)

    def go_up():
        nonlocal current_dir
        current_dir = os.path.dirname(current_dir)
        update_listbox(current_dir)

    def on_ok():
        selected_folders = [
            os.path.join(current_dir, listbox.get(i)) for i in listbox.curselection()
        ]
        root.destroy()
        minerals_result = []
        elements_result = []

        for folder_path in selected_folders:
            folder_name = os.path.basename(folder_path)
            m_start, m_end = extract_meters(folder_name)
            if m_start and m_end:
                minerals, elements = process_files_in_folder(
                    folder_path, m_start, m_end
                )
                if minerals:
                    minerals_result.append(minerals)
                if elements:
                    elements_result.append(elements)

        minerals_result.sort(key=lambda x: (x["From"], x["To"]))
        elements_result.sort(key=lambda x: (x["From"], x["To"]))

        with pd.ExcelWriter("Entregable Proyecto.xlsx", engine="openpyxl") as writer:
            minerals_df = pd.DataFrame(minerals_result)
            elements_df = pd.DataFrame(elements_result)

            # Ordenar columnas alfabéticamente
            minerals_df = minerals_df[
                ["Hole ID", "From", "To"] + sorted(minerals_df.columns[3:])
            ]
            elements_df = elements_df[
                ["Hole ID", "From", "To"] + sorted(elements_df.columns[3:])
            ]

            # Reemplazar NaN con 0.000 solo si no se encontraron archivos
            if minerals_df.isna().any().any():
                minerals_df.fillna(0.000, inplace=True)
            if elements_df.isna().any().any():
                elements_df.fillna(0.000, inplace=True)

            # Excluir columna "TOTAL" si existe
            if "TOTAL" in minerals_df.columns:
                minerals_df.drop(columns=["TOTAL"], inplace=True)
            if "TOTAL" in elements_df.columns:
                elements_df.drop(columns=["TOTAL"], inplace=True)

            # Eliminar columnas con todos los valores 0
            minerals_df = minerals_df.loc[:, (minerals_df != 0).any(axis=0)]
            elements_df = elements_df.loc[:, (elements_df != 0).any(axis=0)]

            minerals_df.to_excel(
                writer, sheet_name="Mineralogía", index=False, startrow=0, header=True
            )
            elements_df.to_excel(
                writer, sheet_name="Geoquímica", index=False, startrow=0, header=True
            )
            for ws_name in writer.sheets:
                format_worksheet(writer.sheets[ws_name])

    up_button = Button(root, text="Atrás", command=go_up)
    up_button.grid(row=1, column=0, pady=10)
    open_button = Button(root, text="Abrir", command=change_directory)
    open_button.grid(row=1, column=1, pady=10)
    ok_button = Button(root, text="Cargar", command=on_ok)
    ok_button.grid(row=1, column=2, pady=10)
    update_listbox(current_dir)
    root.mainloop()


if __name__ == "__main__":
    main()