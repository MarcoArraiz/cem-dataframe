import os
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormatObject, DataBar, Rule
import tkinter as tk
from tkinter import Listbox, Scrollbar, EXTENDED, Button


# Lista de minerales específicos ordenados alfabéticamente
SPECIFIC_MINERALS = sorted([
    "Hematite", "Hematite-Cu", "Goethite", "Goethite-Cu", "Fe-Hidroxide", "Jarosite", 
    "Pyrolusite", "Black-Copper", "Copper-Pitch", "Copper-Wad", "Illmenite", "Magnetite"
    
])

# Lista de minerales secundarios ordenados alfabéticamente
SECONDARY_MINERALS = sorted([
    "Chrysocolla", "Atacamite", "Brochantite", "Malachite", "Chenevixite", "Azurite", "Malachite-Azurite",
    

])

# Lista de minerales terciarios ordenados alfabéticamente
TERTIARY_MINERALS = sorted([
    "Sulphur", "Chalcocite", "Chalcopyrite", "Bornite", "Pyrite", "Barite", 
    "Tennantite-Enargite", "Tetrahedrite", "Molybdenite", "Sphalerite", 
    "Galena", "Stromeyerite", "AgCu"
])

# Lista de elementos químicos (abreviaturas) ordenados alfabéticamente
ELEMENTS = sorted([
    "Cu", "S", "Fe", "Zn", "Mn", "Ag", "Mo", "Pb", "Ba", "As", "Ti", "Al"
])

# Colores de minerales
mineral_colors = {
    "Actinolite": "#E2EFDA",
    "Albite": "#CCCC00",
    "Alunite": "#00FFCC",
    "Andesine": "#C65911",
    "Andradite": "#9999FF",
    "Anhydrite": "#FF66CC",
    "Ankerite": "#3366CC",
    "Amphibole": "#8C941E",
    "Barite": "#666699",
    "Biotite": "#984807",
    "Calcite": "#66FFFF",
    "Apatite": "#3333FF",
    "Carbonate-fluorapatite": "#33CCCC",
    "Chlorite": "#548235",
    "Clay": "#9BC2E6",
    "Clay_Ca": "#9BC2E6",
    "Clay_Mg": "#9BC2E6",
    "Delafossite": "#DC47E7",
    "Diopside": "#00AA7F",
    "Dolomite": "#0099FF",
    "Dravite": "#9933FF",
    "Epidote": "#00FF00",
    "Gree-Gray Sericite": "#F8CBAD",
    "Gypsum": "#FFFFCC",
    "Halloysite": "#66FFFF",
    "Illite": "#8DADF5",
    "Ilmenite": "#808080",
    "Indeterminate Clay": "#203764",
    "Kaolinite": "#99CCFF",
    "K-Feldespar": "#CB2198",
    "Montmorillonite": "#B9CDE5",
    "Muscovite": "#FFFF00",
    "NoData": "#FFFFFF",
    "NoData2": "#FFFFFF",
    "NoData3": "#FFFFFF",
    "Orthoclase": "#CB2198",
    "Plagioclase": "#A50021",
    "Plagioclase Ca": "#A50021",
    "Phengite": "#F4B084",
    "Quartz": "#FFC000",
    "Qz-Fe": "#FFD966",
    "Qz-Ser": "#FF9900",
    "Rutile": "#757171",
    "Sericite": "#DFDA00",
    "Sericite Ti": "#DFDA00",
    "SiNa": "#BF8F00",
    "Smithsonite": "#5F5F5F",
    "Smectite": "#65CDA9",
    "Titanite": "#C0C0C0",
    "Tremolite": "#92D050",
    "Wollastonita": "#CAAFFF",
    "AgCu": "#BFBFBF",
    "Atacamite": "#00FFFF",
    "Black-Copper": "#003300",
    "Bornite": "#FF66CC",
    "Brochantite": "#269960",
    "Chalcocite": "#0066FF",
    "Chalcopyrite": "#FF9933",
    "Chenevixite": "#C3E900",
    "Chrysocolla": "#53C9C7",
    "Copper-pitch": "#AEAAAA",
    "Copper-wad": "#757171",
    "Fe-Hidroxide": "#FFF2CC",
    "Galena": "#8EA9DB",
    "Goethite": "#996633",
    "Goethite-Cu": "#CC9900",
    "Hematite": "#996600",
    "Jarosite": "#663300",
    "Magnetite": "#000000",
    "Malachite": "#339933",
    "Molybdenite": "#8EA9DB",
    "Pyrite": "#FFFF00",
    "Pyrolusite": "#C0C0C0",
    "Sphalerite": "#666699",
    "Stromeyerite": "#FF0000",
    "Tennantite": "#7030A0",
    "Enargite": "#7030A0",
    "Tetrahedrite": "#9966FF",
    "Rutilo": "#747474",
    "Ferrimolybdite": "#8EA9DB",
    "Turquoise": "#9CE0DE",
    "Malachite-Azurite": "#339933",
    # Agrega más minerales y colores aquí si es necesario
}

# Convertir colores hexadecimales a formato aRGB
def hex_to_argb(hex_color):
    return 'FF' + hex_color[1:]

for mineral, color in mineral_colors.items():
    mineral_colors[mineral] = hex_to_argb(color)

def natural_sort_key(s, _nsre=re.compile("([0-9]+)")):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(_nsre, s)]

def extract_meters(folder_name):
    m_start, m_end = folder_name.split(" - ")
    return m_start.rstrip("m").strip(), m_end.rstrip("m").strip()

def process_files_in_folder(folder, m_start, m_end):
    minerals_row = {"Hole ID": "", "From": int(float(m_start)), "To": int(float(m_end))}
    elements_row = {"Hole ID": "", "From": int(float(m_start)), "To": int(float(m_end))}
    processed_files = False

    for file in os.listdir(folder):
        if "Summary" in file and file.endswith(".csv"):
            df = pd.read_csv(os.path.join(folder, file))
            if not df.empty:
                minerals_row.update(process_mineral_data(df))
                elements_row.update(process_element_data(df))
                processed_files = True

    if not processed_files:
        minerals_row.update({col: None for col in minerals_row if col not in ["Hole ID", "From", "To"]})
        elements_row.update({col: None for col in elements_row if col not in ["Hole ID", "From", "To"]})

    return minerals_row, elements_row

def process_mineral_data(df):
    # Limpiar los nombres de las columnas
    df.columns = df.columns.str.strip()
    
    row = {}
    total_presence = df["Presence (%)"].sum()

    if total_presence > 0:
        mineral_presence = (
            df.groupby(df["Class"].str.extract(r"(\D+)")[0].str.strip())["Presence (%)"]
            .sum()
            .to_dict()
        )
        # Fusionar "Goethite" y "Goethite-Cu" bajo "Goethite"
        if "Goethite-Cu" in mineral_presence:
            if "Goethite" in mineral_presence:
                mineral_presence["Goethite"] += mineral_presence.pop("Goethite-Cu")
            else:
                mineral_presence["Goethite"] = mineral_presence.pop("Goethite-Cu")

        # Filtrar solo los minerales específicos, secundarios y terciarios
        row.update({k: v for k, v in mineral_presence.items() if k in SPECIFIC_MINERALS or k in SECONDARY_MINERALS or k in TERTIARY_MINERALS})
        # Incluir todos los demás minerales que no están en los grupos específicos, secundarios o terciarios
        other_minerals = {k: v for k, v in mineral_presence.items() if k not in SPECIFIC_MINERALS and k not in SECONDARY_MINERALS and k not in TERTIARY_MINERALS and k not in ELEMENTS}
        row.update(other_minerals)

    return row

def process_element_data(df):
    row = {}
    element_data = df[df["Class"] == "TOTAL"]

    if not element_data.empty:
        for col in element_data.columns:
            if col not in ["Class", "Presence (%)", "Density", "Hardness"] and col in ELEMENTS:
                value = element_data.iloc[0][col]
                if value > 0:
                    row[col] = value

    return row

from openpyxl.formatting.rule import DataBarRule
def format_worksheet(ws, combined_df, existing_elements, mineral_colors):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Colores de fondo para los diferentes tipos de datos
    primary_fill_color = PatternFill(start_color=hex_to_argb("#F4B084"), end_color=hex_to_argb("#F4B084"), fill_type="solid")
    secondary_fill_color = PatternFill(start_color=hex_to_argb("#C6E0B4"), end_color=hex_to_argb("#C6E0B4"), fill_type="solid")
    tertiary_fill_color = PatternFill(start_color=hex_to_argb("#FFE699"), end_color=hex_to_argb("#FFE699"), fill_type="solid")
    element_fill_color = PatternFill(start_color=hex_to_argb("#FFC9FF"), end_color=hex_to_argb("#FFC9FF"), fill_type="solid")
    excluded_fill_color = PatternFill(start_color=hex_to_argb("#FFFFFF"), end_color=hex_to_argb("#FFFFFF"), fill_type="solid")

    # Aplicar formato a los encabezados y ajustar nombres para PPM
    for cell in ws[1]:  # ws[1] es la fila de encabezados
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.value in SPECIFIC_MINERALS:
            cell.fill = primary_fill_color
        elif cell.value in SECONDARY_MINERALS:
            cell.fill = secondary_fill_color
        elif cell.value in TERTIARY_MINERALS:
            cell.fill = tertiary_fill_color
        elif cell.value in ELEMENTS:
            if cell.value == "Cu":
                cell.fill = element_fill_color
            else:
                col_letter = get_column_letter(cell.col_idx)
                values = [ws[f"{col_letter}{row}"].value for row in range(2, ws.max_row + 1)]
                num_values = len([v for v in values if v is not None])
                num_greater_than_100 = len([v for v in values if v and v > 0.100])
                
                # Evaluar si más del 50% de los valores son mayores que 0.100
                if num_greater_than_100 / num_values <= 0.50:
                    cell.value += " (PPM)"
                    for row in range(2, ws.max_row + 1):
                        cell_to_update = ws[f"{col_letter}{row}"]
                        if cell_to_update.value is not None:
                            cell_to_update.value *= 10000  # Convertir a PPM
                            cell_to_update.number_format = "0"  # Sin decimales
                cell.fill = element_fill_color
        else:
            cell.fill = excluded_fill_color

    # Formato para el resto de las celdas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column >= 4 and " (PPM)" not in ws.cell(row=1, column=cell.column).value:
                cell.number_format = "0.000"

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10 if col < 4 else 16

    # Quitar formato condicional de las celdas con valor 0 o 0.000
    for col in range(4, ws.max_column + 1):
        col_letter = get_column_letter(col)
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.value == 0 or cell.value == 0.000:
                cell.number_format = "0"  # Dejar la celda en 0 sin decimales

    # Aplicar barras de datos condicionales para valores mayores a 0.000
    for col in range(4, ws.max_column + 1):
        column_name = ws.cell(row=1, column=col).value
        if column_name.split(" ")[0] in mineral_colors or column_name.split(" ")[0] in ELEMENTS:
            color = mineral_colors.get(column_name.split(" ")[0], 'FFFF0000')  # Default to red if not found
            col_letter = get_column_letter(col)
            
            # Obtener el valor máximo de la columna después de las conversiones
            values = [cell.value for cell in ws[col_letter][1:] if cell.value is not None and cell.value > 0.000]
            max_val = max(values) if values else 0
            
            # Aplicar barra de datos solo si el valor es mayor que 0.000
            if max_val > 0:
                data_bar_rule = DataBarRule(start_type='num', start_value=0.0001, end_type='num', end_value=max_val, color=color, showValue="None")
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", data_bar_rule)

    # Forzar el color sólido en las barras de datos
    for rule in ws.conditional_formatting._cf_rules:
        if hasattr(rule, 'dataBar') and rule.dataBar:
            rule.dataBar.gradient = False
            rule.dataBar.solidFill = True  # Establecer el relleno sólido

def main():
    root = tk.Tk()
    root.title("Folder Selection")
    current_dir = os.path.abspath(".")
    listbox = Listbox(root, selectmode=EXTENDED, width=100, height=20)
    listbox.grid(row=0, column=0, columnspan=2, padx=10, pady=10)
    scrollbar = Scrollbar(root, orient="vertical", command=listbox.yview)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.grid(row=0, column=2, sticky="ns")

    def update_listbox(directory):
        listbox.delete(0, tk.END)
        items = sorted(
            [item for item in os.listdir(directory) if os.path.isdir(os.path.join(directory, item))],
            key=natural_sort_key,
        )
        for item in items:
            listbox.insert(tk.END, item)

    def change_directory():
        if listbox.curselection():
            nonlocal current_dir
            current_dir = os.path.join(current_dir, listbox.get(listbox.curselection()[0]))
            update_listbox(current_dir)

    def go_up():
        nonlocal current_dir
        current_dir = os.path.dirname(current_dir)
        update_listbox(current_dir)

    def on_ok():
        selected_folders = [
            os.path.join(current_dir, listbox.get(i)) for i in listbox.curselection()
        ]
        root.destroy()
        combined_result = []

        for folder_path in selected_folders:
            folder_name = os.path.basename(folder_path)
            m_start, m_end = extract_meters(folder_name)
            if m_start and m_end:
                minerals, elements = process_files_in_folder(folder_path, m_start, m_end)
                if minerals or elements:
                    combined_row = {**minerals, **elements}
                    combined_result.append(combined_row)

        # Convertir la lista de resultados combinados en un DataFrame
        combined_df = pd.DataFrame(combined_result)

        # Eliminar la columna "TOTAL" si existe
        if "TOTAL" in combined_df.columns:
            combined_df.drop(columns=["TOTAL"], inplace=True)

        # Verificar y filtrar solo las columnas de los minerales específicos, secundarios, terciarios y elementos químicos que están presentes en el DataFrame
        existing_primary_minerals = [mineral for mineral in SPECIFIC_MINERALS if mineral in combined_df.columns]
        existing_secondary_minerals = [mineral for mineral in SECONDARY_MINERALS if mineral in combined_df.columns]
        existing_tertiary_minerals = [mineral for mineral in TERTIARY_MINERALS if mineral in combined_df.columns]
        existing_elements = [col for col in combined_df.columns if col in ELEMENTS]
        excluded_minerals = [col for col in combined_df.columns if col not in ["Hole ID", "From", "To"] + existing_primary_minerals + existing_secondary_minerals + existing_tertiary_minerals + existing_elements]
        
        # Reorganizar las columnas: primero los específicos, luego secundarios, luego terciarios, luego elementos, luego excluidos
        specific_columns = ["Hole ID", "From", "To"] + existing_primary_minerals + existing_secondary_minerals + existing_tertiary_minerals + existing_elements + excluded_minerals
        combined_df = combined_df[specific_columns]

        # Reemplazar NaN con 0.000
        combined_df.fillna(0.000, inplace=True)

        # Ordenar el DataFrame por 'From' y 'To'
        combined_df.sort_values(by=["From", "To"], inplace=True)

        # Guardar el DataFrame en un archivo Excel
        with pd.ExcelWriter("Entregable Proyecto 2.xlsx", engine="openpyxl") as writer:
            combined_df.to_excel(
                writer,
                sheet_name="Mineralogía y Geoquímica",
                index=False,
                startrow=0,
                header=True,
            )
            format_worksheet(writer.sheets["Mineralogía y Geoquímica"], combined_df, existing_elements, mineral_colors)

        print("Archivo Excel 'Entregable Proyecto 2.xlsx' generado exitosamente.")

    Button(root, text="Atrás", command=go_up).grid(row=1, column=0, pady=10)
    Button(root, text="Abrir", command=change_directory).grid(row=1, column=1, pady=10)
    Button(root, text="Cargar", command=on_ok).grid(row=1, column=2, pady=10)

    update_listbox(current_dir)
    root.mainloop()

if __name__ == "__main__":
    main()
